import csv, os, re
import openpyxl
import shutil #automates the process of copying and removal of files and directories.

os.chdir(r'D:\amFOSS_standby\ch-16')
path=os.getcwd()
os.makedirs('csvFiles',exist_ok=True)
newPath= path+ '//' + 'csvFiles'

for excelFiles in os.listdir('.'):
    if not excelFiles.endswith('xlsx'):
        continue
    workbook= openpyxl.load_workbook(excelFiles)

    for sheets in workbook.sheetnames:
        wbName=re.sub('.xlsx', '',excelFiles)
        csvName=wbName + '_' + sheets + '.csv'
        csvFile = open(csvName, 'w', newline='')
        csvWriter = csv.writer(csvFile)
        sheet = workbook.active

    for rowNum in range (1, (sheet.max_row)+1):
        rowData=[]
        for colNum in range (1, (sheet.max_col)+1):
            cellData= sheet.cell(row=rowNum, column=colNum).value
            rowData.append(cellData)
        
        csvWriter=csv.writer(rowData)
        csvFile.close()
        shutil.move(os.path.join(path, csvName), os.path.join(newPath, csvName))
print('Succesfully converted.')

