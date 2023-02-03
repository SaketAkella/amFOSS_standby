import openpyxl

from openpyxl.utils import get_column_letter
wb = openpyxl.Workbook()
tables = 4
sheet = wb.active

for z in range(2, tables + 2):
    sheet['A' + str(z)] = z - 1
    col = get_column_letter(z)
    sheet[col + str(1)] = z - 1

for colm in range(2, tables + 2):
    for row in range(2, tables + 2):
        colNum = get_column_letter(colm)
        sheet[colNum + str(row)] = (colm - 1) * (row - 1)

wb.save('Multiplication Table.xlsx')